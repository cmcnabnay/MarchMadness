import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import Workbook

# Load the Excel file
data = pd.read_excel(
    "C:/Users/cmcna/Git Repositories/MarchMadness/March Madness Statistics.xlsx",
    header=None
).to_numpy()

def build_matrices(start, finish):
    num_rows = finish - start + 1
    A = np.zeros((num_rows, 54))
    B = np.zeros((num_rows, 1))

    for row in range(start, finish + 1, 2):
        row_data = data[row - 1, [3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,24,26,28,30,32,34,36]]
        row_data_2 = data[row, [3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,24,26,28,30,32,34,36]]

        idx1 = row - start
        idx2 = idx1 + 1

        A[idx1, 0:27] = row_data
        A[idx1, 27:54] = row_data_2
        A[idx2, 0:27] = row_data_2
        A[idx2, 27:54] = row_data

    for row in range(start, finish + 1):
        idx = row - start
        B[idx, 0] = data[row - 1, 38]

    return A, B

# Define matchups and ranges
matchups = {
    "8_9": (9, 104),
    "5_12": (107, 202),
    "4_13": (205, 300),
    "6_11": (303, 398),
    "3_14": (401, 496),
    "7_10": (499, 592),
    "2_15": (595, 690),
    "S1": (693, 788),
    "S2": (791, 886),
    "S3": (889, 984),
    "S4": (987, 1082),
    "SS1": (1085, 1180),
    "SS2": (1183, 1278),
    "EE": (1281, 1376),
    "FF": (1379, 1450),
}

# Store matrices
matrices = {}

# Build individual matchups
for key, (start, finish) in matchups.items():
    A, B = build_matrices(start, finish)
    matrices[f"A_{key}"] = A
    matrices[f"B_{key}"] = B

# Combine for each round
Af = np.vstack([matrices[f"A_{k}"] for k in ["8_9", "5_12", "4_13", "6_11", "3_14", "7_10", "2_15"]])
Bf = np.vstack([matrices[f"B_{k}"] for k in ["8_9", "5_12", "4_13", "6_11", "3_14", "7_10", "2_15"]])
matrices["A_first_round"] = Af
matrices["B_first_round"] = Bf

As = np.vstack([matrices[f"A_{k}"] for k in ["S1", "S2", "S3", "S4"]])
Bs = np.vstack([matrices[f"B_{k}"] for k in ["S1", "S2", "S3", "S4"]])
matrices["A_second_round"] = As
matrices["B_second_round"] = Bs

A_ss = np.vstack([matrices["A_SS1"], matrices["A_SS2"]])
B_ss = np.vstack([matrices["B_SS1"], matrices["B_SS2"]])
matrices["A_sweet_sixteen"] = A_ss
matrices["B_sweet_sixteen"] = B_ss

matrices["A_elite_eight"] = matrices["A_EE"]
matrices["B_elite_eight"] = matrices["B_EE"]

matrices["A_final_four"] = matrices["A_FF"]
matrices["B_final_four"] = matrices["B_FF"]

# Combine all rounds
A_all = np.vstack([Af, As, A_ss, matrices["A_EE"], matrices["A_FF"]])
B_all = np.vstack([Bf, Bs, B_ss, matrices["B_EE"], matrices["B_FF"]])
matrices["A_all"] = A_all
matrices["B_all"] = B_all

def full_regression(X, y):
    X_const = sm.add_constant(X)
    y_flat = y.flatten()
    model = sm.OLS(y_flat, X_const).fit()
    return model

# Define all regression names in order
regression_names = [
    "All",
    "8_9",
    "5_12", 
    "4_13",
    "6_11",
    "3_14",
    "7_10",
    "2_15",
    "First Round",
    "Second Round",
    "Sweet Sixteen",
    "Elite Eight",
    "Second Round 1 (S1)",
    "Second Round 2 (S2)",
    "Second Round 3 (S3)",
    "Second Round 4 (S4)",
    "Sweet Sixteen 1 (SS1)",
    "Sweet Sixteen 2 (SS2)",
    "Final Four/Champ (FF)"
]

# Map regression names to matrix keys
key_mapping = {
    "All": "all",
    "8_9": "8_9",
    "5_12": "5_12",
    "4_13": "4_13",
    "6_11": "6_11",
    "3_14": "3_14",
    "7_10": "7_10",
    "2_15": "2_15",
    "First Round": "first_round",
    "Second Round": "second_round",
    "Sweet Sixteen": "sweet_sixteen",
    "Elite Eight": "elite_eight",
    "Second Round 1 (S1)": "S1",
    "Second Round 2 (S2)": "S2",
    "Second Round 3 (S3)": "S3",
    "Second Round 4 (S4)": "S4",
    "Sweet Sixteen 1 (SS1)": "SS1",
    "Sweet Sixteen 2 (SS2)": "SS2",
    "Final Four/Champ (FF)": "FF"
}
print("\n==============================")
print("A_all MATRIX")
print("==============================")
print(A_all)
  

# Column mapping - which variables go in which columns
# Column B-AI, where each column has two variables (row 1 and row 2)
column_structure = [
    ['x1', 'x28'],   # B
    ['x2', 'x29'],   # C
    ['x3', 'x30'],   # D
    ['x4', 'x31'],   # E
    ['x5', 'x32'],   # F
    ['x6', 'x33'],   # G
    ['x7', 'x34'],   # H
    ['x8', 'x35'],   # I
    ['x9', 'x36'],   # J
    ['x10', 'x37'],  # K
    ['x11', 'x38'],  # L
    ['x12', 'x39'],  # M
    ['x13', 'x40'],  # N
    ['x14', 'x41'],  # O
    ['x15', 'x42'],  # P
    ['x16', 'x43'],  # Q
    ['x17', 'x44'],  # R
    ['x18', 'x45'],  # S
    ['x19', 'x46'],  # T
    ['x20', 'x47'],  # U
    ['0', '0'],      # V
    ['x21', 'x48'],  # W
    ['0', '0'],      # X
    ['x22', 'x49'],  # Y
    ['0', '0'],      # Z
    ['x23', 'x50'],  # AA
    ['0', '0'],      # AB
    ['x24', 'x51'],  # AC
    ['0', '0'],      # AD
    ['x25', 'x52'],  # AE
    ['0', '0'],      # AF
    ['x26', 'x53'],  # AG
    ['0', '0'],      # AH
    ['x27', 'x54']   # AI
]

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Regression Coefficients"

# Header row
ws.cell(row=1, column=1, value="Regression")
for i in range(54):
    ws.cell(row=1, column=i+2, value=f"x{i+1}")

current_row = 2

# Run regressions and export coefficients
for reg_name in regression_names:

    key = key_mapping[reg_name]

    print(f"Running regression for {reg_name}")

    A = matrices[f"A_{key}"]
    B = matrices[f"B_{key}"]

    model = full_regression(A, B)

    params = model.params

    # Write regression name
    ws.cell(row=current_row, column=1, value=reg_name)

    # Write coefficients x1 → x54 in columns B → BC
    for i in range(54):

        coeff = params[i+1]   # skip intercept

        ws.cell(row=current_row, column=i+2, value=coeff)

    current_row += 1


# Save Excel file
output_file = "C:/Users/cmcna/Git Repositories/MarchMadness/MarchRegressionCoefficients.xlsx"
wb.save(output_file)

print("\nAll regressions completed.")
print(f"Excel file saved to: {output_file}")